"""Export NPV chart data and editable Excel bar charts."""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Mapping

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

from cement.cement_npv_deterministic import calculate_deterministic_cement_results
from cement.cement_npv_monte_carlo import (
    DEFAULT_RANDOM_SEED as CEMENT_RANDOM_SEED,
    DEFAULT_RETROFIT_BAU_MODE,
    DEFAULT_SAMPLE_SIZE as CEMENT_SAMPLE_SIZE,
    simulate_cement_results,
)
from cement.cement_npv_summary_figures import (
    CEMENT_TECHNOLOGY_LABELS,
    calculate_cement_npv_rankings_from_results,
    cement_npv_distribution_summary,
)
from electricity.electricity_npv_deterministic import (
    calculate_deterministic_electricity_results,
)
from electricity.electricity_npv_monte_carlo import (
    DEFAULT_RANDOM_SEED as ELECTRICITY_RANDOM_SEED,
    DEFAULT_SAMPLE_SIZE as ELECTRICITY_SAMPLE_SIZE,
    simulate_electricity_results,
)
from electricity.electricity_npv_summary_figures import (
    ELECTRICITY_TECHNOLOGY_LABELS,
    calculate_electricity_npv_rankings_from_results,
    electricity_npv_distribution_summary,
)
from npv_summary import deterministic_metric
from npv_summary_plots import fixed_npv_bar_axis_config


OUTPUT_DIR = Path(__file__).resolve().parents[1] / "Excel plots"


def _distribution_stat(
    summary: Mapping[str, Mapping[str, float]],
    statistic: str,
) -> dict[str, float]:
    return {label: values[statistic] for label, values in summary.items()}


def _axis_limits_for_sheet(
    sector: str,
    npv_scale: str,
    summary: Mapping[str, Mapping[str, float]],
    deterministic_values: Mapping[str, float],
) -> tuple[tuple[float, float], tuple[float, ...]]:
    return fixed_npv_bar_axis_config(
        sector=sector,
        npv_scale=npv_scale,
        distribution_summary=summary,
        deterministic_values=deterministic_values,
    )


def _axis_major_unit(sector: str, npv_scale: str) -> float:
    sector_key = sector.lower()
    if sector_key == "cement" and npv_scale == "MEUR":
        return 200.0
    if sector_key == "electricity" and npv_scale == "MEUR":
        return 1000.0
    if npv_scale in {"EUR/t", "EUR/MWh"}:
        return 50.0
    return 0.0


def _append_table(
    ws,
    start_row: int,
    title: str,
    values: Mapping[str, float],
    extra_columns: Mapping[str, Mapping[str, float]] | None = None,
) -> tuple[int, int, int]:
    """Append one sorted chart-data table and return data coordinates."""

    extra_columns = extra_columns or {}
    sorted_items = sorted(values.items(), key=lambda item: item[1], reverse=True)
    headers = ["Technology", title, *extra_columns]
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True)
    header_row = start_row + 1
    for column, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=column, value=header).font = Font(bold=True)
    for row_offset, (label, value) in enumerate(sorted_items, start=1):
        row = header_row + row_offset
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=float(value))
        for column_offset, column_name in enumerate(extra_columns, start=3):
            ws.cell(
                row=row,
                column=column_offset,
                value=float(extra_columns[column_name][label]),
            )

    first_data_row = header_row + 1
    last_data_row = header_row + len(sorted_items)
    return header_row, first_data_row, last_data_row


def _add_bar_chart(
    ws,
    title: str,
    header_row: int,
    first_data_row: int,
    last_data_row: int,
    anchor: str,
    axis_title: str,
    axis_limits: tuple[float, float],
    axis_major_unit: float,
    reverse_categories: bool = True,
) -> None:
    """Add one editable Excel horizontal bar chart."""

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = "Technology"
    chart.x_axis.title = axis_title
    chart.x_axis.scaling.min = axis_limits[0]
    chart.x_axis.scaling.max = axis_limits[1]
    if axis_major_unit > 0:
        chart.x_axis.majorUnit = axis_major_unit
    chart.y_axis.scaling.orientation = "maxMin" if reverse_categories else "minMax"
    chart.height = 9
    chart.width = 16
    chart.legend = None
    data = Reference(ws, min_col=2, min_row=header_row, max_row=last_data_row)
    categories = Reference(ws, min_col=1, min_row=first_data_row, max_row=last_data_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, anchor)


def _add_sheet(
    wb: Workbook,
    sheet_name: str,
    sector: str,
    npv_scale: str,
    axis_title: str,
    summary: Mapping[str, Mapping[str, float]],
    deterministic_values: Mapping[str, float],
) -> None:
    ws = wb.create_sheet(sheet_name)
    axis_limits, axis_ticks = _axis_limits_for_sheet(
        sector,
        npv_scale,
        summary,
        deterministic_values,
    )
    axis_major_unit = _axis_major_unit(sector, npv_scale)
    ws["A1"] = "Shared x-axis minimum"
    ws["B1"] = axis_limits[0]
    ws["A2"] = "Shared x-axis maximum"
    ws["B2"] = axis_limits[1]
    ws["A3"] = "Readable PNG tick labels"
    ws["B3"] = ", ".join(f"{tick:g}" for tick in axis_ticks)
    ws["A4"] = "Suggested Excel major unit"
    ws["B4"] = axis_major_unit
    ws["A5"] = "Use these limits for both charts on this sheet."
    for cell in ("A1", "A2", "A3", "A4", "A5"):
        ws[cell].font = Font(bold=True)

    mean_values = _distribution_stat(summary, "mean")
    header_row, first_row, last_row = _append_table(
        ws,
        start_row=6,
        title="Monte Carlo mean",
        values=mean_values,
        extra_columns={
            "Median": _distribution_stat(summary, "median"),
            "P05": _distribution_stat(summary, "p05"),
            "P95": _distribution_stat(summary, "p95"),
        },
    )
    _add_bar_chart(
        ws,
        title="Monte Carlo mean NPV",
        header_row=header_row,
        first_data_row=first_row,
        last_data_row=last_row,
        anchor="G5",
        axis_title=axis_title,
        axis_limits=axis_limits,
        axis_major_unit=axis_major_unit,
    )

    header_row, first_row, last_row = _append_table(
        ws,
        start_row=21,
        title="Deterministic",
        values=deterministic_values,
    )
    _add_bar_chart(
        ws,
        title="Deterministic NPV",
        header_row=header_row,
        first_data_row=first_row,
        last_data_row=last_row,
        anchor="G22",
        axis_title=axis_title,
        axis_limits=axis_limits,
        axis_major_unit=axis_major_unit,
    )

    for column in ("A", "B", "C", "D", "E"):
        ws.column_dimensions[column].width = 22


def _ranking_display_summary(ranking_summary, labels: Mapping[str, str]):
    return ranking_summary.assign(
        display_label=ranking_summary["technology"].map(labels).fillna(
            ranking_summary["technology"]
        )
    ).sort_values(["average_rank", "technology"])


def _add_ranking_sheet(
    wb: Workbook,
    sheet_name: str,
    ranking_summary,
    labels: Mapping[str, str],
) -> None:
    ws = wb.create_sheet(sheet_name)
    display_summary = _ranking_display_summary(ranking_summary, labels)
    headers = ["Technology", "Average rank", "Probability rank 1", "Probability top 3"]
    for column, header in enumerate(headers, start=1):
        ws.cell(row=1, column=column, value=header).font = Font(bold=True)
    for row_offset, row in enumerate(display_summary.itertuples(index=False), start=2):
        ws.cell(row=row_offset, column=1, value=row.display_label)
        ws.cell(row=row_offset, column=2, value=float(row.average_rank))
        ws.cell(row=row_offset, column=3, value=float(row.probability_rank_1))
        ws.cell(row=row_offset, column=4, value=float(row.probability_top_3))

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Specific NPV average ranking"
    chart.y_axis.title = "Technology"
    chart.x_axis.title = "Average rank (1 = highest NPV)"
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = len(display_summary)
    chart.y_axis.scaling.orientation = "maxMin"
    chart.height = 11
    chart.width = 18
    chart.legend = None
    data = Reference(ws, min_col=2, min_row=1, max_row=len(display_summary) + 1)
    categories = Reference(ws, min_col=1, min_row=2, max_row=len(display_summary) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, "F2")

    for column in ("A", "B", "C", "D"):
        ws.column_dimensions[column].width = 24


def _cement_deterministic_values(npv_scale: str) -> dict[str, float]:
    metric_column = "npv_eur" if npv_scale == "MEUR" else "npv_eur_per_t"
    scale = 1_000_000.0 if npv_scale == "MEUR" else 1.0
    return deterministic_metric(
        results_by_item=calculate_deterministic_cement_results(),
        labels=CEMENT_TECHNOLOGY_LABELS,
        metric_column=metric_column,
        scale=scale,
    )


def _electricity_deterministic_values(npv_scale: str) -> dict[str, float]:
    metric_column = "npv_eur" if npv_scale == "MEUR" else "npv_eur_per_mwh"
    scale = 1_000_000.0 if npv_scale == "MEUR" else 1.0
    return deterministic_metric(
        results_by_item=calculate_deterministic_electricity_results(),
        labels=ELECTRICITY_TECHNOLOGY_LABELS,
        metric_column=metric_column,
        scale=scale,
    )


def save_workbook(output_dir: Path = OUTPUT_DIR, run_date: date | None = None) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    output_date = run_date or date.today()
    output_path = output_dir / f"{output_date.isoformat()}-NPV_chart_data.xlsx"

    cement_results = simulate_cement_results(
        sample_size=CEMENT_SAMPLE_SIZE,
        random_seed=CEMENT_RANDOM_SEED,
        retrofit_bau_mode=DEFAULT_RETROFIT_BAU_MODE,
    )
    electricity_results = simulate_electricity_results(
        sample_size=ELECTRICITY_SAMPLE_SIZE,
        random_seed=ELECTRICITY_RANDOM_SEED,
    )

    wb = Workbook()
    wb.remove(wb.active)
    _add_sheet(
        wb,
        "Cement MEUR",
        "cement",
        "MEUR",
        "NPV (million EUR)",
        cement_npv_distribution_summary(cement_results, npv_scale="MEUR"),
        _cement_deterministic_values("MEUR"),
    )
    _add_sheet(
        wb,
        "Cement EUR per t",
        "cement",
        "EUR/t",
        "NPV (EUR/t)",
        cement_npv_distribution_summary(cement_results, npv_scale="EUR/t"),
        _cement_deterministic_values("EUR/t"),
    )
    _add_sheet(
        wb,
        "Electricity MEUR",
        "electricity",
        "MEUR",
        "NPV (million EUR)",
        electricity_npv_distribution_summary(electricity_results, npv_scale="MEUR"),
        _electricity_deterministic_values("MEUR"),
    )
    _add_sheet(
        wb,
        "Electricity EUR per MWh",
        "electricity",
        "EUR/MWh",
        "NPV (EUR/MWh)",
        electricity_npv_distribution_summary(electricity_results, npv_scale="EUR/MWh"),
        _electricity_deterministic_values("EUR/MWh"),
    )
    _, cement_specific_ranking = calculate_cement_npv_rankings_from_results(
        cement_results,
        npv_scale="EUR/t",
    )
    _add_ranking_sheet(
        wb,
        "Cement specific ranking",
        cement_specific_ranking,
        CEMENT_TECHNOLOGY_LABELS,
    )
    _, electricity_specific_ranking = calculate_electricity_npv_rankings_from_results(
        electricity_results,
        npv_scale="EUR/MWh",
    )
    _add_ranking_sheet(
        wb,
        "Electricity specific ranking",
        electricity_specific_ranking,
        ELECTRICITY_TECHNOLOGY_LABELS,
    )
    wb.save(output_path)
    return output_path


def main() -> None:
    print(save_workbook())


if __name__ == "__main__":
    main()
